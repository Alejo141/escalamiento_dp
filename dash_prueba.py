"""
Dashboard SAC
- Fuente: OneDrive via Microsoft Graph API
- Roles: viewer (público) / admin (login requerido para editar)
- Credenciales en Streamlit Secrets
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import requests
from io import BytesIO
from datetime import datetime
import numpy as np

# PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ─────────────────────────────────────────────
# CONFIGURACIÓN GENERAL
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Dashboard SAC",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────

SHEET_ABIERTOS      = "Consolidado"
SHEET_CERRADOS      = "Gestion_SAC"
SHEET_LOG           = "Log_Cambios"
BINS_SEMAFORO       = [-999, 5, 10, 999]
LABELS_SEMAFORO     = ["🟢 En tiempo", "🟡 En riesgo", "🔴 Vencido"]
COLUMNAS_REQUERIDAS = ["FechaCreacion", "Responsable", "NombreSeccionales", "NUI"]

# Paleta ampliada — usada en todos los gráficos de categorías
PALETA = [
    "#8f5cda","#7069d8","#3a81d5","#38a9d2","#4cb2ca",
    "#a78bfa","#60a5fa","#34d399","#f472b6","#fb923c",
    "#facc15","#e879f9","#2dd4bf","#818cf8","#f87171",
    "#4ade80","#38bdf8","#c084fc","#fb7185","#a3e635",
]

# Columnas visibles en las tablas del dashboard (en orden)
COLUMNAS_TABLA = [
    "NUI", "NombreSeccionales", "Id_Tickets", "Semaforo",
    "Menu", "SubMenu1", "SubMenu2", "SubMenu3", "FechaCreacion", "Dias_Habiles",
    "Creador_gestion", "Responsable", "Fecha Asignación",
    "Descripción", "Fecha Respuesta",
]

# ─────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────

st.markdown("""
<style>
.main, .block-container { background-color: #0f1116; }
section[data-testid="stSidebar"] { background-color: #161a24; }
h1, h2, h3, h4 { color: #8f5cda !important; }
p, label, .stMarkdown { color: #d1d5db; }
div[data-testid="metric-container"] {
    background: linear-gradient(135deg, #161a24, #1e2130);
    border: 1px solid #2a2f3a;
    border-left: 4px solid #8f5cda;
    padding: 14px 18px; border-radius: 12px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.3);
}
div[data-testid="stMetricValue"] { color: #ffffff !important; font-size: 2rem !important; }
div[data-testid="stMetricLabel"] { color: #9ca3af !important; }
div[data-testid="stMetricDelta"] { font-size: 0.85rem; }
.stDataFrame { border-radius: 10px; overflow: hidden; }
.stButton > button {
    background: linear-gradient(135deg, #8f5cda, #3a81d5);
    color: white; border: none; border-radius: 8px;
    padding: 8px 20px; font-weight: 600; transition: opacity .2s;
}
.stButton > button:hover { opacity: .85; }
hr { border-color: #2a2f3a; }
.badge {
    display: inline-block; background: #1e2130;
    border: 1px solid #3a81d5; color: #38a9d2;
    border-radius: 20px; padding: 2px 14px;
    font-size: 0.75rem; margin-bottom: 12px;
}
.role-admin {
    display: inline-block; background: #1a2e1a;
    border: 1px solid #4ade80; color: #4ade80;
    border-radius: 20px; padding: 2px 12px; font-size: 0.75rem;
}
.role-viewer {
    display: inline-block; background: #1e2130;
    border: 1px solid #6b7280; color: #9ca3af;
    border-radius: 20px; padding: 2px 12px; font-size: 0.75rem;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# MICROSOFT GRAPH — TOKEN
# ─────────────────────────────────────────────

def get_token() -> str:
    """
    Obtiene access token usando refresh_token (flujo delegado).
    El refresh_token se renueva automáticamente en cada llamada.
    """
    cfg  = st.secrets["microsoft"]
    resp = requests.post(
        f"https://login.microsoftonline.com/{cfg['tenant_id']}/oauth2/v2.0/token",
        data={
            "grant_type":    "refresh_token",
            "client_id":     cfg["client_id"],
            "client_secret": cfg["client_secret"],
            "refresh_token": cfg["refresh_token"],
            "scope":         "Files.ReadWrite offline_access User.Read",
        },
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def headers() -> dict:
    return {"Authorization": f"Bearer {get_token()}"}

# ─────────────────────────────────────────────
# MICROSOFT GRAPH — LEER / ESCRIBIR
# ─────────────────────────────────────────────

def _drive_url() -> str:
    """Construye la URL base del archivo en el OneDrive del usuario autenticado."""
    file_path = st.secrets["microsoft"]["file_path"]
    return f"https://graph.microsoft.com/v1.0/me/drive/root:{file_path}:"


# Lista fija de responsables — usada también en la normalización al cargar
# Tabla de unificación: variantes conocidas → nombre canónico
# Solo corrige errores tipográficos o variaciones del mismo nombre
UNIFICACION_RESPONSABLES = {
    # DIANA - JUAN SEBASTIAN
    "DIANA SANCHEZ - JUAN MORA":                    "DIANA - JUAN SEBASTIAN",
    "DIANA- JUAN SEBASTIAN":                        "DIANA - JUAN SEBASTIAN",
    "DIANA -JUAN SEBASTIAN":                        "DIANA - JUAN SEBASTIAN",
    "DIANAJUAN SEBASTIAN":                          "DIANA - JUAN SEBASTIAN",
    # DIANA - GERALDIN
    "DIANA- GERALDIN":                              "DIANA - GERALDIN",
    "DIANA -GERALDIN":                              "DIANA - GERALDIN",
    "DIANA GERALDIN":                               "DIANA - GERALDIN",
    # DIANA - GERALDIN - JURIDICO
    "DIANA- GERALDIN - JURIDICO":                   "DIANA - GERALDIN - JURIDICO",
    "DIANA GERALDIN JURIDICO":                      "DIANA - GERALDIN - JURIDICO",
    "DIANA-GERALDIN-JURIDICO":                      "DIANA - GERALDIN - JURIDICO",
    # JAVIER PRADA - GERALDIN
    "JAVIER PRADA- GERALDIN":                       "JAVIER PRADA - GERALDIN",
    "JAVIER PRADA -GERALDIN":                       "JAVIER PRADA - GERALDIN",
    "JAVIERPRADA - GERALDIN":                       "JAVIER PRADA - GERALDIN",
    "JAVIER PRADA GERALDIN":                        "JAVIER PRADA - GERALDIN",
    "JAVIER PRADA - GERALDIN VARGAS":               "JAVIER PRADA - GERALDIN",
    "JAVIER PRADA- GERALDIN VARGAS":                "JAVIER PRADA - GERALDIN",
    # JAVIER PRADA
    "JAVIER  PRADA":                                "JAVIER PRADA",
    "JAVIE PRADA":                                  "JAVIER PRADA",
    # JURIDICO - JAVIER PRADA
    "JURIDICO- JAVIER PRADA":                       "JURIDICO - JAVIER PRADA",
    "JURIDICO -JAVIER PRADA":                       "JURIDICO - JAVIER PRADA",
    "JURIDICOJAVIER PRADA":                         "JURIDICO - JAVIER PRADA",
    "JURIDICO JAVIER PRADA":                        "JURIDICO - JAVIER PRADA",
    "JURIDICO- JAVIER PRADA":                       "JURIDICO - JAVIER PRADA",
    # GERALDIN - SAC
    "GERALDIN- SAC":                                "GERALDIN - SAC",
    "GERALDIN -SAC":                                "GERALDIN - SAC",
    "GERALDINSAC":                                  "GERALDIN - SAC",
    # GERALDIN - JURIDICO
    "GERALDIN- JURIDICO":                           "GERALDIN - JURIDICO",
    "GERALDIN JURIDICO":                            "GERALDIN - JURIDICO",
    # GERALDIN - SUI
    "GERALDIN- SUI":                                "GERALDIN - SUI",
    "GERALDIN SUI":                                 "GERALDIN - SUI",
    # GESTORES SOCIALES
    "GESTORESSOCIALES":                             "GESTORES SOCIALES",
    "GESTORES  SOCIALES":                           "GESTORES SOCIALES",
    "GESTOR SOCIAL":                                "GESTORES SOCIALES",
    # DIANA - JUAN SEBASTIAN variantes con apellidos
    "DIANA SANCHEZ- JUAN MORA":                     "DIANA - JUAN SEBASTIAN",
    "DIANA SANCHEZ JUAN MORA":                      "DIANA - JUAN SEBASTIAN",
    # NO SE ESCALA
    "NO SE ESCALA":                                 "NO SE ESCALA",
    "NOESCALA":                                     "NO SE ESCALA",
    "NO ESCALA":                                    "NO SE ESCALA",
}

def _normalizar_col_responsable(df: pd.DataFrame) -> pd.DataFrame:
    """
    Unifica variantes del mismo responsable usando:
    1. Tabla de unificación exacta (insensible a mayúsculas y espacios extra)
    2. Si no hay match exacto, conserva el valor original del Excel
    """
    unif_upper = {k.upper().strip(): v for k, v in UNIFICACION_RESPONSABLES.items()}

    def _match(valor):
        if pd.isna(valor) or str(valor).strip() == "":
            return valor
        v = str(valor).strip().upper()
        # Normalizar espacios múltiples
        import re
        v_clean = re.sub(' +', ' ', v)
        if v_clean in unif_upper:
            return unif_upper[v_clean]
        return str(valor).strip()  # conservar original sin modificar

    if "Responsable" in df.columns:
        df = df.copy()
        df["Responsable"] = df["Responsable"].apply(_match)
    return df


@st.cache_data(ttl=60, show_spinner="📥 Descargando datos desde OneDrive…")
def cargar_datos() -> tuple[pd.DataFrame, pd.DataFrame]:
    url  = _drive_url() + "/content"
    resp = requests.get(url, headers=headers(), timeout=30)
    resp.raise_for_status()
    buf  = BytesIO(resp.content)
    df_a = pd.read_excel(buf, sheet_name=SHEET_ABIERTOS, engine="openpyxl")
    buf.seek(0)
    df_c = pd.read_excel(buf, sheet_name=SHEET_CERRADOS, engine="openpyxl")
    # Normalizar responsables en ambas hojas
    df_a = _normalizar_col_responsable(df_a)
    df_c = _normalizar_col_responsable(df_c)
    return df_a, df_c


def registrar_log(
    usuario: str,
    accion: str,       # "EDITAR" | "ELIMINAR"
    nui: str,
    campo: str,
    valor_anterior: str,
    valor_nuevo: str,
) -> None:
    """
    Agrega una fila al log de auditoría en la hoja Log_Cambios del Excel.
    Si la hoja no existe la crea.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    url_content = _drive_url() + "/content"
    resp_get    = requests.get(url_content, headers=headers(), timeout=30)
    resp_get.raise_for_status()
    wb = load_workbook(BytesIO(resp_get.content))

    # Crear hoja de log si no existe
    if SHEET_LOG not in wb.sheetnames:
        ws_log = wb.create_sheet(SHEET_LOG)
        ws_log.append(["Fecha", "Usuario", "Acción", "NUI", "Campo", "Valor anterior", "Valor nuevo"])
    else:
        ws_log = wb[SHEET_LOG]

    ws_log.append([
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        usuario,
        accion,
        str(nui),
        campo,
        str(valor_anterior),
        str(valor_nuevo),
    ])

    buf_out = BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)
    requests.put(url_content, headers={**headers(),
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        data=buf_out.read(), timeout=60).raise_for_status()


def guardar_datos(df_abiertos: pd.DataFrame, df_cerrados: pd.DataFrame) -> None:
    """
    Descarga el archivo original, actualiza los datos fila por fila
    preservando el formato de tabla (Table) de cada hoja, y lo sube.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    # 1 — Descargar el archivo original para preservar formato
    url_content = _drive_url() + "/content"
    resp_get    = requests.get(url_content, headers=headers(), timeout=30)
    resp_get.raise_for_status()
    buf_orig = BytesIO(resp_get.content)

    wb = load_workbook(buf_orig)

    def actualizar_hoja(wb, nombre_hoja: str, df: pd.DataFrame) -> None:
        ws = wb[nombre_hoja]

        # Detectar la tabla Excel en la hoja
        tabla = list(ws.tables.values())

        # Limpiar datos existentes (mantener fila de encabezado)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # Escribir nuevos datos fila por fila
        for r_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
            for c_idx, value in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Actualizar el rango de la tabla para que incluya los nuevos datos
        if tabla:
            t = tabla[0]
            ref_start = t.ref.split(":")[0]  # ej: A1
            col_end   = get_column_letter(len(df.columns))
            row_end   = len(df) + 1          # +1 por el encabezado
            t.ref     = f"{ref_start}:{col_end}{row_end}"

    actualizar_hoja(wb, SHEET_ABIERTOS, df_abiertos)
    actualizar_hoja(wb, SHEET_CERRADOS, df_cerrados)

    # 2 — Guardar en buffer y subir a OneDrive
    buf_out = BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)

    resp_put = requests.put(
        url_content,
        headers={
            **headers(),
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
        data=buf_out.read(),
        timeout=60,
    )
    resp_put.raise_for_status()

# ─────────────────────────────────────────────
# AUTENTICACIÓN — ROLES
# ─────────────────────────────────────────────

def verificar_credenciales(usuario: str, password: str) -> str | None:
    usuarios = st.secrets.get("usuarios", {})
    cfg      = usuarios.get(usuario)
    if cfg and cfg.get("password") == password:
        return cfg.get("rol")
    return None


def es_admin() -> bool:
    return st.session_state.get("rol") == "admin"


def mostrar_login_sidebar() -> None:
    with st.sidebar:
        st.divider()
        if es_admin():
            st.markdown(
                f'<div class="role-admin">🔐 Admin: {st.session_state["usuario"]}</div>',
                unsafe_allow_html=True,
            )
            if st.button("🔌 Cerrar sesión", use_container_width=True):
                st.session_state.pop("rol",     None)
                st.session_state.pop("usuario", None)
                st.rerun()
        else:
            # Login oculto — se despliega al hacer clic en la flecha
            with st.expander("🔑 ›", expanded=False):
                with st.form("form_login", clear_on_submit=True):
                    usr = st.text_input("Usuario")
                    pwd = st.text_input("Contraseña", type="password")
                    ok  = st.form_submit_button("Iniciar sesión", use_container_width=True)
                if ok:
                    rol = verificar_credenciales(usr, pwd)
                    if rol:
                        st.session_state["rol"]     = rol
                        st.session_state["usuario"] = usr
                        st.rerun()
                    else:
                        st.error("Usuario o contraseña incorrectos.")

# ─────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────

# Festivos colombianos fijos (Ley 51/1983 + Ley 270/1996)
# Se actualiza anualmente agregando el año correspondiente
def _festivos_colombia() -> list:
    festivos = []
    for anio in range(2020, 2031):
        # Festivos fijos
        fijos = [
            f"{anio}-01-01",  # Año Nuevo
            f"{anio}-05-01",  # Día del Trabajo
            f"{anio}-07-20",  # Independencia
            f"{anio}-08-07",  # Batalla de Boyacá
            f"{anio}-12-08",  # Inmaculada Concepción
            f"{anio}-12-25",  # Navidad
        ]
        festivos.extend(fijos)

        # Festivos "Ley Emiliani" — se mueven al lunes siguiente si no caen en lunes
        def lunes_siguiente(fecha_str):
            d = pd.Timestamp(fecha_str)
            if d.weekday() == 0:
                return d
            dias = 7 - d.weekday()
            return d + pd.Timedelta(days=dias)

        emiliani = [
            f"{anio}-01-06",  # Reyes Magos
            f"{anio}-03-19",  # San José
            f"{anio}-06-29",  # San Pedro y San Pablo
            f"{anio}-08-15",  # Asunción
            f"{anio}-10-12",  # Día de la Raza
            f"{anio}-11-01",  # Todos los Santos
            f"{anio}-11-11",  # Independencia Cartagena
        ]
        for f in emiliani:
            festivos.append(str(lunes_siguiente(f).date()))

        # Semana Santa (Jueves y Viernes Santo) — cálculo algoritmo de Gauss
        a = anio % 19
        b = anio // 100
        c = anio % 100
        d = b // 4
        e = b % 4
        f = (b + 8) // 25
        g = (b - f + 1) // 3
        h = (19*a + b - d - g + 15) % 30
        i = c // 4
        k = c % 4
        l = (32 + 2*e + 2*i - h - k) % 7
        m = (a + 11*h + 22*l) // 451
        mes  = (h + l - 7*m + 114) // 31
        dia  = ((h + l - 7*m + 114) % 31) + 1
        pascua   = pd.Timestamp(year=anio, month=mes, day=dia)
        jueves   = pascua - pd.Timedelta(days=3)
        viernes  = pascua - pd.Timedelta(days=2)
        ascension = lunes_siguiente(str((pascua + pd.Timedelta(days=39)).date()))
        corpus    = lunes_siguiente(str((pascua + pd.Timedelta(days=60)).date()))
        sagrado   = lunes_siguiente(str((pascua + pd.Timedelta(days=68)).date()))
        festivos.extend([
            str(jueves.date()), str(viernes.date()),
            str(ascension.date()), str(corpus.date()), str(sagrado.date()),
        ])

    return [np.datetime64(f, "D") for f in festivos]

FESTIVOS_CO = np.array(_festivos_colombia(), dtype="datetime64[D]")


def calcular_dias_habiles(df: pd.DataFrame) -> pd.Series:
    """Días hábiles entre FechaCreacion y hoy, excluyendo festivos colombianos."""
    hoy    = np.datetime64(pd.Timestamp.today().normalize(), "D")
    fechas = pd.to_datetime(df["FechaCreacion"], errors="coerce").values.astype("datetime64[D]")
    valido = ~pd.isnull(df["FechaCreacion"])
    dias   = np.where(
        valido,
        np.busday_count(fechas, hoy, holidays=FESTIVOS_CO),
        np.nan,
    )
    return pd.Series(dias, index=df.index)


def aplicar_semaforo(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega Dias_Habiles y Semaforo_KPI (para métricas internas).
    La columna Semaforo original del Excel NO se toca — se muestra tal cual en tablas.
    """
    df = df.copy()
    df["FechaCreacion"] = pd.to_datetime(df["FechaCreacion"], errors="coerce")
    df["Dias_Habiles"]  = calcular_dias_habiles(df)
    df["Semaforo_KPI"]  = pd.cut(df["Dias_Habiles"], bins=BINS_SEMAFORO, labels=LABELS_SEMAFORO)
    return df


def a_excel(df: pd.DataFrame) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False)
    return buf.getvalue()


def generar_pdf(
    titulo: str,
    kpis: dict,
    filtros_activos: dict,
    figs: list,          # lista de (nombre, fig plotly)
    df_tabla: pd.DataFrame,
) -> bytes:
    """
    Genera un reporte PDF con:
    - Encabezado con título, fecha y filtros activos
    - Tarjetas de KPIs
    - Gráficas exportadas como imágenes
    - Tabla resumen (primeras 50 filas)
    """
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        title=titulo,
    )

    # ── Estilos ──
    estilos = getSampleStyleSheet()
    st_titulo = ParagraphStyle("Titulo", parent=estilos["Title"],
                                fontSize=18, textColor=colors.HexColor("#8f5cda"),
                                spaceAfter=6, alignment=TA_LEFT)
    st_sub    = ParagraphStyle("Sub", parent=estilos["Heading2"],
                                fontSize=11, textColor=colors.HexColor("#3a81d5"),
                                spaceBefore=12, spaceAfter=4)
    st_normal = ParagraphStyle("Normal2", parent=estilos["Normal"],
                                fontSize=8, textColor=colors.HexColor("#374151"))
    st_footer = ParagraphStyle("Footer", parent=estilos["Normal"],
                                fontSize=7, textColor=colors.HexColor("#9ca3af"),
                                alignment=TA_RIGHT)
    st_kpi_val = ParagraphStyle("KpiVal", parent=estilos["Normal"],
                                 fontSize=22, textColor=colors.HexColor("#ffffff"),
                                 alignment=TA_CENTER, fontName="Helvetica-Bold")
    st_kpi_lbl = ParagraphStyle("KpiLbl", parent=estilos["Normal"],
                                 fontSize=7, textColor=colors.HexColor("#9ca3af"),
                                 alignment=TA_CENTER)

    story = []
    ancho_pagina = landscape(A4)[0] - 3*cm  # ancho útil

    # ── Encabezado ──
    story.append(Paragraph(titulo, st_titulo))
    story.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
        f"Registros en vista: <b>{len(df_tabla)}</b>",
        st_normal,
    ))

    # Filtros activos
    filtros_str = "  |  ".join(f"{k}: <b>{v}</b>" for k, v in filtros_activos.items() if v)
    if filtros_str:
        story.append(Paragraph(f"Filtros activos: {filtros_str}", st_normal))

    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#8f5cda"), spaceAfter=10))

    # ── KPIs ──
    story.append(Paragraph("Indicadores clave", st_sub))
    col_ancho = ancho_pagina / max(len(kpis), 1)
    kpi_data = [[
        Paragraph(str(v), st_kpi_val) for v in kpis.values()
    ], [
        Paragraph(k, st_kpi_lbl) for k in kpis.keys()
    ]]
    kpi_table = Table(kpi_data, colWidths=[col_ancho] * len(kpis))
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#161a24")),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",   (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 8),
        ("LINEAFTER", (0, 0), (-2, -1), 0.5, colors.HexColor("#2a2f3a")),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#8f5cda")),
        ("ROUNDEDCORNERS", [6]),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 14))

    # ── Gráficas ──
    if figs:
        story.append(Paragraph("Gráficas", st_sub))
        # Renderizar en pares (2 por fila)
        pares = [figs[i:i+2] for i in range(0, len(figs), 2)]
        img_ancho = (ancho_pagina - 1*cm) / 2
        img_alto  = img_ancho * 0.55

        for par in pares:
            fila_imgs = []
            for nombre, fig in par:
                try:
                    img_bytes = fig.to_image(format="png", width=900, height=500,
                                              scale=1.5, engine="kaleido")
                    img_buf = BytesIO(img_bytes)
                    rl_img  = RLImage(img_buf, width=img_ancho, height=img_alto)
                    fila_imgs.append(rl_img)
                except Exception:
                    fila_imgs.append(Paragraph(f"[{nombre}: imagen no disponible]", st_normal))

            if len(fila_imgs) == 1:
                fila_imgs.append(Spacer(img_ancho, img_alto))

            t = Table([fila_imgs], colWidths=[img_ancho, img_ancho])
            t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
            story.append(t)
            story.append(Spacer(1, 6))

    # ── Resumen de casos por resolver ──
    story.append(Paragraph("Resumen de casos por resolver", st_sub))

    d = df_tabla.copy()

    # ── 1. Por Responsable ──
    story.append(Paragraph("<b>Casos por Responsable</b>", st_normal))
    if "Responsable" in d.columns:
        df_resp = (d.groupby("Responsable").size()
                    .reset_index(name="Tickets")
                    .sort_values("Tickets", ascending=False))
        filas_resp = [["Responsable", "Tickets"]] + df_resp.fillna("").astype(str).values.tolist()
        t_resp = Table(filas_resp, colWidths=[ancho_pagina * 0.7, ancho_pagina * 0.3])
        t_resp.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#8f5cda")),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#1e2130"), colors.HexColor("#161a24")]),
            ("TEXTCOLOR",     (0, 1), (-1, -1), colors.HexColor("#d1d5db")),
            ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
            ("ALIGN",         (1, 0), (1,  -1), "CENTER"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#2a2f3a")),
        ]))
        story.append(t_resp)
        story.append(Spacer(1, 10))

    # ── 2. Por Seccional ──
    story.append(Paragraph("<b>Casos por Seccional</b>", st_normal))
    if "NombreSeccionales" in d.columns:
        df_secc = (d.groupby("NombreSeccionales").size()
                    .reset_index(name="Tickets")
                    .sort_values("Tickets", ascending=False))
        filas_secc = [["Seccional", "Tickets"]] + df_secc.fillna("").astype(str).values.tolist()
        t_secc = Table(filas_secc, colWidths=[ancho_pagina * 0.7, ancho_pagina * 0.3])
        t_secc.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#3a81d5")),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#1e2130"), colors.HexColor("#161a24")]),
            ("TEXTCOLOR",     (0, 1), (-1, -1), colors.HexColor("#d1d5db")),
            ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
            ("ALIGN",         (1, 0), (1,  -1), "CENTER"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#2a2f3a")),
        ]))
        story.append(t_secc)
        story.append(Spacer(1, 10))

    # ── 3. Por Semáforo ──
    story.append(Paragraph("<b>Casos por Estado (Semáforo)</b>", st_normal))
    if "Semaforo_KPI" in d.columns:
        df_sem = (d.groupby("Semaforo_KPI").size()
                   .reset_index(name="Tickets")
                   .sort_values("Tickets", ascending=False))
        filas_sem = [["Estado", "Tickets"]] + df_sem.fillna("").astype(str).values.tolist()
        t_sem = Table(filas_sem, colWidths=[ancho_pagina * 0.7, ancho_pagina * 0.3])
        t_sem.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#374151")),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#1e2130"), colors.HexColor("#161a24")]),
            ("TEXTCOLOR",     (0, 1), (-1, -1), colors.HexColor("#d1d5db")),
            ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
            ("ALIGN",         (1, 0), (1,  -1), "CENTER"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#2a2f3a")),
        ]))
        story.append(t_sem)
        story.append(Spacer(1, 10))

    # ── 4. Por Categoría (SubMenu1) ──
    story.append(Paragraph("<b>Casos por Categoría (SubMenu1)</b>", st_normal))
    if "SubMenu1" in d.columns:
        df_sub1 = (d.groupby("SubMenu1").size()
                    .reset_index(name="Tickets")
                    .sort_values("Tickets", ascending=False))
        filas_sub1 = [["Categoría", "Tickets"]] + df_sub1.fillna("").astype(str).values.tolist()
        t_sub1 = Table(filas_sub1, colWidths=[ancho_pagina * 0.7, ancho_pagina * 0.3])
        t_sub1.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#7069d8")),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#1e2130"), colors.HexColor("#161a24")]),
            ("TEXTCOLOR",     (0, 1), (-1, -1), colors.HexColor("#d1d5db")),
            ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
            ("ALIGN",         (1, 0), (1,  -1), "CENTER"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#2a2f3a")),
        ]))
        story.append(t_sub1)
        story.append(Spacer(1, 10))

    # ── 5. Top 10 casos más antiguos (mayor días hábiles) ──
    story.append(Paragraph("<b>Top 10 casos más antiguos sin resolver</b>", st_normal))
    cols_antiguos = [c for c in ["NUI", "NombreSeccionales", "Responsable", "SubMenu1",
                                  "FechaCreacion", "Dias_Habiles", "Descripción"] if c in d.columns]
    if cols_antiguos and "Dias_Habiles" in d.columns:
        df_ant = (d[cols_antiguos].sort_values("Dias_Habiles", ascending=False).head(10).copy())
        if "FechaCreacion" in df_ant.columns:
            df_ant["FechaCreacion"] = pd.to_datetime(df_ant["FechaCreacion"], errors="coerce").dt.strftime("%d/%m/%Y").fillna("")
        # Truncar Descripción para que no explote la tabla
        if "Descripción" in df_ant.columns:
            df_ant["Descripción"] = df_ant["Descripción"].astype(str).str[:60]
        col_w_ant = ancho_pagina / len(cols_antiguos)
        filas_ant = [cols_antiguos] + df_ant.fillna("").astype(str).values.tolist()
        t_ant = Table(filas_ant, colWidths=[col_w_ant] * len(cols_antiguos), repeatRows=1)
        t_ant.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#f87171")),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#2e1a1a"), colors.HexColor("#1e2130")]),
            ("TEXTCOLOR",     (0, 1), (-1, -1), colors.HexColor("#fca5a5")),
            ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 4),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#7f1d1d")),
        ]))
        story.append(t_ant)

    # ── Pie de página ──
    story.append(Spacer(1, 12))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.HexColor("#2a2f3a"), spaceAfter=4))
    story.append(Paragraph(
        f"Dashboard SAC — Reporte automático · {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        st_footer,
    ))

    doc.build(story)
    return buf.getvalue()


def validar_columnas(df: pd.DataFrame, nombre: str) -> None:
    faltantes = [c for c in COLUMNAS_REQUERIDAS if c not in df.columns]
    if faltantes:
        st.warning(f"⚠️ **{nombre}** le faltan columnas: `{', '.join(faltantes)}`")

# ─────────────────────────────────────────────
# CARGA DE DATOS
# ─────────────────────────────────────────────

try:
    df_abiertos, df_cerrados = cargar_datos()
except requests.HTTPError as e:
    st.error(f"❌ Error al conectar con OneDrive: `{e.response.status_code} {e.response.reason}`")
    st.info("""
**Posibles causas:**
- `401 Unauthorized` → client_id, client_secret o tenant_id incorrectos
- `403 Forbidden` → permisos no otorgados o falta "Grant admin consent"
- `404 Not Found` → file_path o user_id incorrecto
    """)
    st.stop()
except Exception as e:
    st.error(f"❌ Error inesperado: `{e}`")
    st.stop()

validar_columnas(df_abiertos, SHEET_ABIERTOS)
validar_columnas(df_cerrados, SHEET_CERRADOS)

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────

mostrar_login_sidebar()

with st.sidebar:
    st.markdown("## ⚡ Filtros")

# ─────────────────────────────────────────────
# NAVEGACIÓN
# ─────────────────────────────────────────────

TABS = ["⚡ Escalamiento", "💡 Casos Gestionados", "🔧 Gestionar"] if es_admin() \
       else ["⚡ Escalamiento", "💡 Casos Gestionados"]

dashboard = st.selectbox("Seleccionar módulo", TABS, label_visibility="collapsed")

rol_html = '<span class="role-admin">🔐 Admin</span>' if es_admin() \
           else '<span class="role-viewer">👁️ Viewer</span>'

st.markdown("""
<style>
.btn-verde > button {
    background: linear-gradient(135deg, #16a34a, #15803d) !important;
    color: white !important; border: none !important;
    border-radius: 8px !important; font-weight: 600 !important;
}
.btn-verde > button:hover { opacity: .85 !important; }
</style>
""", unsafe_allow_html=True)

col_badge, col_btn = st.columns([7, 1])
with col_badge:
    st.markdown(
        f'<div class="badge">🔌 OneDrive · {datetime.now().strftime("%d/%m/%Y %H:%M")} &nbsp;{rol_html}</div>',
        unsafe_allow_html=True,
    )
with col_btn:
    st.markdown('<div class="btn-verde">', unsafe_allow_html=True)
    if st.button("⚡ Actualizar", use_container_width=True):
        st.cache_data.clear()
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
# ═══════════════════════════════════════════════
# MÓDULO: GESTIONAR  (solo admin)
# ═══════════════════════════════════════════════

if dashboard == "🔧 Gestionar":

    if not es_admin():
        st.error("🔒 Acceso restringido. Inicia sesión como administrador.")
        st.stop()

    st.title("🔧 Gestión de Tickets")
    df_edit = df_abiertos.copy()

    with st.expander("🔎 Filtros de búsqueda", expanded=True):
        c1, c2, c3, c4 = st.columns(4)
        f_nui         = c1.text_input("NUI")
        f_ticket      = c2.text_input("Id_Tickets")
        f_seccional   = c3.selectbox("Seccional",   ["Todos"] + sorted(df_edit["NombreSeccionales"].dropna().unique().tolist()))
        f_responsable = c4.selectbox("Responsable", ["Todos"] + sorted(df_edit["Responsable"].dropna().unique().tolist()))

    if f_nui:
        df_edit = df_edit[df_edit["NUI"].astype(str).str.contains(f_nui, case=False)]
    if f_ticket and "Id_Tickets" in df_edit.columns:
        df_edit = df_edit[df_edit["Id_Tickets"].astype(str).str.contains(f_ticket, case=False)]
    if f_seccional != "Todos":
        df_edit = df_edit[df_edit["NombreSeccionales"] == f_seccional]
    if f_responsable != "Todos":
        df_edit = df_edit[df_edit["Responsable"] == f_responsable]

    # ── Tabla 1: Gestión SAC ──
    st.subheader("🗂️ Gestión SAC")
    df_cerrados_edit = df_cerrados.copy()
    if f_nui:
        df_cerrados_edit = df_cerrados_edit[df_cerrados_edit["NUI"].astype(str).str.contains(f_nui, case=False)]
    if f_seccional != "Todos" and "NombreSeccionales" in df_cerrados_edit.columns:
        df_cerrados_edit = df_cerrados_edit[df_cerrados_edit["NombreSeccionales"] == f_seccional]
    if f_responsable != "Todos" and "Responsable" in df_cerrados_edit.columns:
        df_cerrados_edit = df_cerrados_edit[df_cerrados_edit["Responsable"] == f_responsable]
    cols_cerrados = [c for c in COLUMNAS_TABLA if c in df_cerrados_edit.columns]
    st.markdown(f"**{len(df_cerrados_edit)}** registros")
    st.dataframe(df_cerrados_edit[cols_cerrados] if cols_cerrados else df_cerrados_edit,
                 use_container_width=True, height=280)

    st.divider()

    # ── Tabla 2: Consolidado ──
    st.subheader("📑 Consolidado")
    st.markdown(f"**{len(df_edit)}** registros encontrados")
    cols_visibles = [c for c in COLUMNAS_TABLA if c in df_edit.columns]
    st.dataframe(df_edit[cols_visibles] if cols_visibles else df_edit,
                 use_container_width=True, height=280)

    if df_edit.empty:
        st.info("No hay registros que coincidan.")
        st.stop()

    st.divider()
    st.subheader("⚙️ Editar registro")

    index_sel = st.selectbox(
        "Selecciona un registro", df_edit.index,
        format_func=lambda i: f"#{i} — NUI: {df_abiertos.loc[i,'NUI']} | Responsable: {df_abiertos.loc[i,'Responsable']}",
    )
    registro = df_abiertos.loc[index_sel]

    # Lista dinámica: valores únicos ya normalizados del Excel
    responsables_lista = sorted(df_abiertos["Responsable"].dropna().unique().tolist())

    # Opciones fijas de SubMenu1
    OPCIONES_SUBMENU1 = [
        "Reposición",
        "Sin SISFV Instalada",
        "Restricciones Acceso",
        "Usuario Ausente",
        "Traslados No Autorizado",
        "Suspension del servicio",
        "Reconexion del servicio",
        "SISFV Funcional",
        "Ausencia de Partes",
        "SISFV Abandonado",
        "Equipos modificados",
    ]

    with st.form("form_edicion"):
        # Campos de solo lectura (informativos)
        col1, col2 = st.columns(2)
        col1.text_input("NUI",      registro.get("NUI", ""),           disabled=True)
        col2.text_input("Semáforo", str(registro.get("Semaforo", "")), disabled=True)

        # Responsable: lista desplegable con los existentes en Consolidado
        responsable_actual = registro.get("Responsable", "")
        idx_resp = responsables_lista.index(responsable_actual) if responsable_actual in responsables_lista else 0
        nuevo_responsable = st.selectbox("Responsable", responsables_lista, index=idx_resp)

        # SubMenu1: lista fija de opciones
        submenu_actual = str(registro.get("SubMenu1", ""))
        idx_sub = OPCIONES_SUBMENU1.index(submenu_actual) if submenu_actual in OPCIONES_SUBMENU1 else 0
        nuevo_submenu1 = st.selectbox("SubMenu1", OPCIONES_SUBMENU1, index=idx_sub)

        # Descripción editable
        nueva_descripcion = st.text_area("Descripción", registro.get("Descripción", ""), height=120)

        col_a, col_b, _  = st.columns([1, 1, 3])
        guardar  = col_a.form_submit_button("✅ Guardar cambios")
        eliminar = col_b.form_submit_button("🗑️ Eliminar registro", type="secondary")

    if guardar:
        usuario_log = st.session_state.get("usuario", "desconocido")
        nui_log     = str(registro.get("NUI", ""))
        # Registrar cambios campo por campo
        cambios = {
            "Responsable": (registro.get("Responsable", ""), nuevo_responsable),
            "SubMenu1":    (registro.get("SubMenu1", ""),    nuevo_submenu1),
            "Descripción": (registro.get("Descripción", ""), nueva_descripcion),
        }
        df_abiertos.loc[index_sel, "Responsable"] = nuevo_responsable
        df_abiertos.loc[index_sel, "SubMenu1"]    = nuevo_submenu1
        df_abiertos.loc[index_sel, "Descripción"] = nueva_descripcion
        try:
            with st.spinner("Guardando en OneDrive…"):
                guardar_datos(df_abiertos, df_cerrados)
                for campo, (ant, nvo) in cambios.items():
                    if str(ant) != str(nvo):
                        registrar_log(usuario_log, "EDITAR", nui_log, campo, ant, nvo)
            st.success("✅ Cambios guardados en OneDrive.")
        except Exception as e:
            st.error(f"❌ Error al guardar: `{e}`")
        st.cache_data.clear()
        st.rerun()

    if eliminar:
        if st.session_state.get("confirm_delete") != index_sel:
            st.session_state["confirm_delete"] = index_sel
            st.warning("⚠️ Presiona **Eliminar registro** de nuevo para confirmar.")
        else:
            nui_elim     = str(registro.get("NUI", ""))
            usuario_elim = st.session_state.get("usuario", "desconocido")
            df_abiertos.drop(index=index_sel, inplace=True)
            try:
                with st.spinner("Guardando cambios…"):
                    guardar_datos(df_abiertos, df_cerrados)
                    registrar_log(usuario_elim, "ELIMINAR", nui_elim, "—", "Registro completo", "Eliminado")
                st.warning("🗑️ Registro eliminado correctamente.")
            except Exception as e:
                st.error(f"❌ Error al eliminar: `{e}`")
            st.session_state.pop("confirm_delete", None)
            st.cache_data.clear()
            st.rerun()

    # ── Log de auditoría (solo admin) ──
    st.divider()
    st.subheader("📋 Registro de auditoría")
    try:
        url_log  = _drive_url() + "/content"
        resp_log = requests.get(url_log, headers=headers(), timeout=30)
        resp_log.raise_for_status()
        df_log = pd.read_excel(BytesIO(resp_log.content),
                               sheet_name=SHEET_LOG, engine="openpyxl")
        df_log = df_log.sort_values("Fecha", ascending=False) if "Fecha" in df_log.columns else df_log
        st.dataframe(df_log, use_container_width=True, height=300)
        st.download_button(
            label="📥 Exportar log a Excel",
            data=a_excel(df_log),
            file_name=f"SAC_log_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="export_log",
        )
    except Exception:
        st.info("Aún no hay registros de auditoría. Se crearán al hacer la primera edición.")

    st.stop()

# ═══════════════════════════════════════════════
# MÓDULO: ESCALAMIENTO / CASOS CERRADOS
# ═══════════════════════════════════════════════

es_escalamiento = dashboard == "⚡ Escalamiento"
df_base = df_abiertos.copy() if es_escalamiento else df_cerrados.copy()
df_base = aplicar_semaforo(df_base)

# ─────────────────────────────────────────────
# FILTROS BIDIRECCIONALES
# Cada filtro muestra solo las opciones que tienen
# datos considerando TODOS los demás filtros activos
# ─────────────────────────────────────────────

def _opts(df_in: pd.DataFrame, col: str) -> list:
    if col not in df_in.columns:
        return []
    return sorted(df_in[col].dropna().unique().tolist())

def _aplicar_sin(df_in, excluir_col,
                 responsables, seccionales, menus,
                 submenus1, submenus2, submenus3) -> pd.DataFrame:
    """Aplica todos los filtros excepto el de excluir_col."""
    d = df_in.copy()
    if excluir_col != "Responsable"       and responsables: d = d[d["Responsable"].isin(responsables)]
    if excluir_col != "NombreSeccionales" and seccionales:  d = d[d["NombreSeccionales"].isin(seccionales)]
    if excluir_col != "Menu"              and menus and "Menu" in d.columns:     d = d[d["Menu"].isin(menus)]
    if excluir_col != "SubMenu1"          and submenus1 and "SubMenu1" in d.columns: d = d[d["SubMenu1"].isin(submenus1)]
    if excluir_col != "SubMenu2"          and submenus2 and "SubMenu2" in d.columns: d = d[d["SubMenu2"].isin(submenus2)]
    if excluir_col != "SubMenu3"          and submenus3 and "SubMenu3" in d.columns: d = d[d["SubMenu3"].isin(submenus3)]
    return d

# Leer selecciones previas de session_state para calcular opciones
_prev_resp  = st.session_state.get("f_resp",  [])
_prev_secc  = st.session_state.get("f_secc",  [])
_prev_menu  = st.session_state.get("f_menu",  [])
_prev_sub1  = st.session_state.get("f_sub1",  [])
_prev_sub2  = st.session_state.get("f_sub2",  [])
_prev_sub3  = st.session_state.get("f_sub3",  [])

with st.sidebar:
    # Opciones de cada filtro = df filtrado por todos los DEMÁS
    opts_resp = _opts(_aplicar_sin(df_base, "Responsable",       _prev_resp, _prev_secc, _prev_menu, _prev_sub1, _prev_sub2, _prev_sub3), "Responsable")
    opts_secc = _opts(_aplicar_sin(df_base, "NombreSeccionales", _prev_resp, _prev_secc, _prev_menu, _prev_sub1, _prev_sub2, _prev_sub3), "NombreSeccionales")
    opts_menu = _opts(_aplicar_sin(df_base, "Menu",              _prev_resp, _prev_secc, _prev_menu, _prev_sub1, _prev_sub2, _prev_sub3), "Menu")
    opts_sub1 = _opts(_aplicar_sin(df_base, "SubMenu1",          _prev_resp, _prev_secc, _prev_menu, _prev_sub1, _prev_sub2, _prev_sub3), "SubMenu1")
    opts_sub2 = _opts(_aplicar_sin(df_base, "SubMenu2",          _prev_resp, _prev_secc, _prev_menu, _prev_sub1, _prev_sub2, _prev_sub3), "SubMenu2")
    opts_sub3 = _opts(_aplicar_sin(df_base, "SubMenu3",          _prev_resp, _prev_secc, _prev_menu, _prev_sub1, _prev_sub2, _prev_sub3), "SubMenu3")

    # Limpiar selecciones previas que ya no están en opciones
    def _clean(prev, opts): return [v for v in prev if v in opts]

    responsables = st.multiselect("Responsable", opts_resp, default=_clean(_prev_resp, opts_resp), key="f_resp")
    seccionales  = st.multiselect("Seccional",   opts_secc, default=_clean(_prev_secc, opts_secc), key="f_secc")
    menus        = st.multiselect("Menú",         opts_menu, default=_clean(_prev_menu, opts_menu), key="f_menu")
    submenus1    = st.multiselect("SubMenu1",     opts_sub1, default=_clean(_prev_sub1, opts_sub1), key="f_sub1")
    submenus2    = st.multiselect("SubMenu2",     opts_sub2, default=_clean(_prev_sub2, opts_sub2), key="f_sub2")
    submenus3    = st.multiselect("SubMenu3",     opts_sub3, default=_clean(_prev_sub3, opts_sub3), key="f_sub3")

    # Rango de fechas
    fecha_min, fecha_max = df_base["FechaCreacion"].min(), df_base["FechaCreacion"].max()
    if pd.notna(fecha_min) and pd.notna(fecha_max) and fecha_min != fecha_max:
        rango_fechas = st.date_input(
            "Rango de fechas",
            value=(fecha_min.date(), fecha_max.date()),
            min_value=fecha_min.date(), max_value=fecha_max.date(),
        )
    else:
        rango_fechas = None

# Aplicar todos los filtros al df final
df = _aplicar_sin(df_base, "__ninguno__",
                  responsables, seccionales, menus,
                  submenus1, submenus2, submenus3)
if rango_fechas and len(rango_fechas) == 2:
    df = df[(df["FechaCreacion"].dt.date >= rango_fechas[0]) &
            (df["FechaCreacion"].dt.date <= rango_fechas[1])]

# KPIs
titulo = "⚡ Gestión Escalamiento" if es_escalamiento else "💡 Gestión Casos Gestionados"
st.title(titulo)

total     = len(df)
secciones = df["NombreSeccionales"].nunique()
nuis      = df["NUI"].nunique()
prom_dias = round(df["Dias_Habiles"].mean(), 1) if not df.empty else 0
vencidos  = (df["Semaforo_KPI"] == "🔴 Vencido").sum()
en_riesgo = (df["Semaforo_KPI"] == "🟡 En riesgo").sum()

k1, k2, k3, k4, k5, k6 = st.columns(6)
k1.metric("Total Tickets", total)
k2.metric("Seccionales",   secciones)
k3.metric("NUIs únicos",   nuis)
k4.metric("Promedio días", prom_dias)
k5.metric("🔴 Vencidos",   vencidos,  delta=f"-{vencidos}"  if vencidos  else None, delta_color="inverse")
k6.metric("🟡 En riesgo",  en_riesgo, delta=f"-{en_riesgo}" if en_riesgo else None, delta_color="inverse")

st.divider()

# ── TABLA PRIMERO ──
st.subheader("⚡ Detalle de tickets")
busqueda_top = st.text_input("🔎 Búsqueda rápida", "", key="busqueda_top")
df_tabla_top = df.copy()
if busqueda_top:
    mask = df_tabla_top.astype(str).apply(lambda col: col.str.contains(busqueda_top, case=False)).any(axis=1)
    df_tabla_top = df_tabla_top[mask]
cols_top = [c for c in COLUMNAS_TABLA if c in df_tabla_top.columns]
df_top   = df_tabla_top[cols_top].copy() if cols_top else df_tabla_top.copy()

# Formatear columnas de fecha a dd/mm/yyyy
COLS_FECHA = ["FechaCreacion", "Fecha Asignación", "Fecha Respuesta"]
for col_f in COLS_FECHA:
    if col_f in df_top.columns:
        df_top[col_f] = pd.to_datetime(df_top[col_f], errors="coerce").dt.strftime("%d/%m/%Y").fillna("")

st.dataframe(
    df_top.style.apply(lambda row: [
        "background-color: #1a2e1a" if row.get("Semaforo_KPI") == "🟢 En tiempo"
        else "background-color: #2e2a1a" if row.get("Semaforo_KPI") == "🟡 En riesgo"
        else "background-color: #2e1a1a" if row.get("Semaforo_KPI") == "🔴 Vencido"
        else "" for _ in row], axis=1),
    use_container_width=True, height=380,
)
cols_export = [c for c in COLUMNAS_TABLA if c in df_tabla_top.columns]

btn_col1, btn_col2 = st.columns([1, 1])

with btn_col1:
    st.download_button(
        label="📥 Exportar tabla a Excel",
        data=a_excel(df_tabla_top[cols_export]),
        file_name=f"SAC_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="export_top",
    )

with btn_col2:
    # ── Construir figuras para el PDF ──
    def _build_figs_pdf(df_in):
        figs_pdf = []
        L = dict(plot_bgcolor="#161a24", paper_bgcolor="#161a24",
                 font=dict(color="#d1d5db"), margin=dict(l=40,r=20,t=40,b=40))
        # Seccional
        if "NombreSeccionales" in df_in.columns:
            d = df_in.groupby("NombreSeccionales").size().reset_index(name="Tickets").sort_values("Tickets", ascending=True)
            f = px.bar(d, x="Tickets", y="NombreSeccionales", orientation="h",
                       color="NombreSeccionales", color_discrete_sequence=PALETA, template="plotly_dark",
                       title="Tickets por Seccional")
            f.update_layout(**L, showlegend=False)
            figs_pdf.append(("Seccional", f))
        # Semáforo
        if "Semaforo_KPI" in df_in.columns:
            d = df_in["Semaforo_KPI"].value_counts().reset_index()
            d.columns = ["Estado","Cantidad"]
            f = px.pie(d, names="Estado", values="Cantidad", hole=0.5,
                       color="Estado", color_discrete_map={"🟢 En tiempo":"#4ade80","🟡 En riesgo":"#facc15","🔴 Vencido":"#f87171"},
                       template="plotly_dark", title="Estado del Servicio")
            f.update_layout(**L)
            figs_pdf.append(("Semáforo", f))
        # SubMenu1
        if "SubMenu1" in df_in.columns:
            d = df_in.groupby("SubMenu1").size().reset_index(name="Tickets").sort_values("Tickets", ascending=True)
            f = px.bar(d, x="Tickets", y="SubMenu1", orientation="h",
                       color="SubMenu1", color_discrete_sequence=PALETA, template="plotly_dark",
                       title="Tickets por Categoría")
            f.update_layout(**L, showlegend=False)
            figs_pdf.append(("SubMenu1", f))
        # Top 10 Responsables
        if "Responsable" in df_in.columns:
            d = df_in.groupby("Responsable").size().reset_index(name="Tickets").sort_values("Tickets",ascending=False).head(10).sort_values("Tickets",ascending=True)
            f = px.bar(d, x="Tickets", y="Responsable", orientation="h",
                       color="Responsable", color_discrete_sequence=PALETA, template="plotly_dark",
                       title="Top 10 Responsables")
            f.update_layout(**L, showlegend=False)
            figs_pdf.append(("Top Responsables", f))
        # Tendencia
        if "FechaCreacion" in df_in.columns and df_in["FechaCreacion"].notna().any():
            d = (df_in.dropna(subset=["FechaCreacion"])
                       .assign(Mes=lambda x: x["FechaCreacion"].dt.to_period("M").astype(str))
                       .groupby("Mes").size().reset_index(name="Tickets"))
            f = px.line(d, x="Mes", y="Tickets", markers=True,
                        template="plotly_dark", color_discrete_sequence=["#a78bfa"],
                        title="Tendencia de Creación")
            f.update_layout(**L)
            figs_pdf.append(("Tendencia", f))
        return figs_pdf

    # Filtros activos para el encabezado del PDF
    filtros_pdf = {}
    if responsables: filtros_pdf["Responsable"] = ", ".join(responsables)
    if seccionales:  filtros_pdf["Seccional"]   = ", ".join(seccionales)
    if menus:        filtros_pdf["Menú"]         = ", ".join(menus)
    if submenus1:    filtros_pdf["SubMenu1"]     = ", ".join(submenus1)
    if submenus2:    filtros_pdf["SubMenu2"]     = ", ".join(submenus2)
    if submenus3:    filtros_pdf["SubMenu3"]     = ", ".join(submenus3)

    kpis_pdf = {
        "Total Tickets": total,
        "Seccionales":   secciones,
        "NUIs únicos":   nuis,
        "Promedio días": prom_dias,
        "Vencidos":      int(vencidos),
        "En riesgo":     int(en_riesgo),
    }

    try:
        pdf_bytes = generar_pdf(
            titulo=titulo,
            kpis=kpis_pdf,
            filtros_activos=filtros_pdf,
            figs=_build_figs_pdf(df),
            df_tabla=df_tabla_top[cols_export],
        )
        st.download_button(
            label="📄 Descargar reporte PDF",
            data=pdf_bytes,
            file_name=f"SAC_reporte_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            mime="application/pdf",
            key="export_pdf",
        )
    except Exception as e:
        st.warning(f"PDF no disponible: `{e}`")

st.divider()

LAYOUT = dict(plot_bgcolor="#161a24", paper_bgcolor="#161a24", margin=dict(l=0,r=0,t=30,b=0),
              legend=dict(bgcolor="rgba(0,0,0,0)", font_color="#d1d5db"),
              font=dict(color="#d1d5db"))

# ── FILA 1: Seccional + Semáforo ──
col_g1, col_g2 = st.columns([2, 1])
with col_g1:
    st.subheader("🏘️ Tickets por Seccional")
    df_sec = df.groupby("NombreSeccionales").size().reset_index(name="Tickets").sort_values("Tickets", ascending=True)
    fig_bar = px.bar(df_sec, x="Tickets", y="NombreSeccionales", orientation="h",
                     color="NombreSeccionales", color_discrete_sequence=PALETA, template="plotly_dark")
    fig_bar.update_layout(**LAYOUT, showlegend=False, yaxis_title=None)
    st.plotly_chart(fig_bar, use_container_width=True, key="fig_bar")

with col_g2:
    st.subheader("🚦 Estado del Servicio")
    df_sem = df["Semaforo_KPI"].value_counts().reset_index()
    df_sem.columns = ["Estado", "Cantidad"]
    fig_pie = px.pie(df_sem, names="Estado", values="Cantidad", hole=0.55, template="plotly_dark",
                     color="Estado", color_discrete_map={"🟢 En tiempo": "#4ade80", "🟡 En riesgo": "#facc15", "🔴 Vencido": "#f87171"})
    fig_pie.update_layout(**{**LAYOUT, "legend": dict(orientation="h", y=-0.15, bgcolor="rgba(0,0,0,0)", font_color="#d1d5db")})
    fig_pie.update_traces(textinfo="percent+label", textfont_size=12)
    st.plotly_chart(fig_pie, use_container_width=True, key="fig_pie")

# ── FILA 3: SubMenu1 — ancho completo ──
st.subheader("⚡ Tickets por Categoría")
if "SubMenu1" in df.columns:
    df_sub = (df.groupby("SubMenu1").size().reset_index(name="Tickets")
                .sort_values("Tickets", ascending=True))
    fig_sub = px.bar(df_sub, x="Tickets", y="SubMenu1", orientation="h",
                     color="SubMenu1", color_discrete_sequence=PALETA, template="plotly_dark")
    fig_sub.update_layout(**LAYOUT, showlegend=False, yaxis_title=None)
    st.plotly_chart(fig_sub, use_container_width=True, key="fig_sub")
else:
    st.info("Columna SubMenu1 no disponible.")

# ── FILA 4: SubMenu2 ← izq · SubMenu3 → der ──
col_g5, col_g6 = st.columns(2)
with col_g5:
    st.subheader("🔌 Tickets por SubCategoría")
    if "SubMenu2" in df.columns and df["SubMenu2"].notna().any():
        df_sub2 = (df.groupby("SubMenu2").size().reset_index(name="Tickets")
                     .sort_values("Tickets", ascending=True))
        fig_sub2 = px.bar(df_sub2, x="Tickets", y="SubMenu2", orientation="h",
                          color="SubMenu2", color_discrete_sequence=PALETA, template="plotly_dark")
        fig_sub2.update_layout(**LAYOUT, showlegend=False, yaxis_title=None)
        st.plotly_chart(fig_sub2, use_container_width=True, key="fig_sub2")
    else:
        st.info("Columna SubMenu2 no disponible o sin datos.")

with col_g6:
    st.subheader("💡 Tickets por Detalle")
    if "SubMenu3" in df.columns and df["SubMenu3"].notna().any():
        df_sub3 = (df.groupby("SubMenu3").size().reset_index(name="Tickets")
                     .sort_values("Tickets", ascending=True))
        fig_sub3 = px.bar(df_sub3, x="Tickets", y="SubMenu3", orientation="h",
                          color="SubMenu3", color_discrete_sequence=PALETA, template="plotly_dark")
        fig_sub3.update_layout(**LAYOUT, showlegend=False, yaxis_title=None)
        st.plotly_chart(fig_sub3, use_container_width=True, key="fig_sub3")
    else:
        st.info("Columna SubMenu3 no disponible o sin datos.")

# ── FILA 5: Top 10 Responsables ← izq · Tiempo promedio cierre → der ──
col_g7, col_g8 = st.columns(2)
with col_g7:
    st.subheader("👷 Top 10 Responsables con más Tickets")
    if "Responsable" in df.columns:
        df_resp = (df.groupby("Responsable").size().reset_index(name="Tickets")
                     .sort_values("Tickets", ascending=False).head(10)
                     .sort_values("Tickets", ascending=True))
        fig_resp = px.bar(df_resp, x="Tickets", y="Responsable", orientation="h",
                          color="Responsable", color_discrete_sequence=PALETA, template="plotly_dark")
        fig_resp.update_layout(**LAYOUT, showlegend=False, yaxis_title=None)
        st.plotly_chart(fig_resp, use_container_width=True, key="fig_resp")

with col_g8:
    st.subheader("⏱️ Tiempo Promedio de Atención por Responsable")
    if "Responsable" in df.columns and df["Dias_Habiles"].notna().any():
        df_cierre = (df.groupby("Responsable")["Dias_Habiles"].mean().reset_index()
                       .rename(columns={"Dias_Habiles": "Dias para Cierre"})
                       .sort_values("Dias para Cierre", ascending=True))
        fig_cierre = px.bar(df_cierre, x="Dias para Cierre", y="Responsable", orientation="h",
                            color="Responsable", color_discrete_sequence=PALETA, template="plotly_dark",
                            labels={"Dias para Cierre": "Días promedio"})
        fig_cierre.update_layout(**LAYOUT, showlegend=False, yaxis_title=None)
        st.plotly_chart(fig_cierre, use_container_width=True, key="fig_cierre")

# ── FILA 2: Tendencia + Treemap ──
col_g3, col_g4 = st.columns(2)
with col_g3:
    st.subheader("📈 Tendencia de creación")
    if df["FechaCreacion"].notna().any():
        df_trend = (df.dropna(subset=["FechaCreacion"])
                      .assign(Mes=lambda x: x["FechaCreacion"].dt.to_period("M").astype(str))
                      .groupby("Mes").size().reset_index(name="Tickets"))
        fig_line = px.line(df_trend, x="Mes", y="Tickets", markers=True,
                           template="plotly_dark", color_discrete_sequence=["#a78bfa"])
        fig_line.update_layout(**LAYOUT)
        fig_line.update_traces(line_width=2.5, marker_size=7, line_color="#a78bfa",
                               marker=dict(color="#f472b6", size=8))
        st.plotly_chart(fig_line, use_container_width=True, key="fig_line")
    else:
        st.info("Sin datos de fecha disponibles.")

with col_g4:
    st.subheader("🗺️ Cobertura Seccional × Responsable")
    df_tree = df.groupby(["NombreSeccionales", "Responsable"]).size().reset_index(name="Tickets")
    fig_tree = px.treemap(df_tree, path=["NombreSeccionales", "Responsable"], values="Tickets",
                          color="Tickets", color_continuous_scale=["#3a81d5", "#8f5cda", "#f472b6"],
                          template="plotly_dark")
    fig_tree.update_layout(**LAYOUT)
    st.plotly_chart(fig_tree, use_container_width=True, key="fig_tree")

# ── FILA 6: Heatmap Seccional × Mes ──
st.divider()
st.subheader("🌡️ Heatmap de tickets por Seccional × Mes")
if "FechaCreacion" in df.columns and "NombreSeccionales" in df.columns and df["FechaCreacion"].notna().any():
    df_heat = (
        df.dropna(subset=["FechaCreacion"])
          .assign(Mes=lambda x: x["FechaCreacion"].dt.to_period("M").astype(str))
          .groupby(["NombreSeccionales", "Mes"])
          .size().reset_index(name="Tickets")
    )
    fig_heat = px.density_heatmap(
        df_heat, x="Mes", y="NombreSeccionales", z="Tickets",
        color_continuous_scale=["#161a24","#3a81d5","#8f5cda","#f472b6"],
        template="plotly_dark",
        labels={"Mes": "Mes", "NombreSeccionales": "Seccional", "Tickets": "Tickets"},
    )
    fig_heat.update_layout(**LAYOUT, coloraxis_colorbar=dict(title="Tickets"))
    st.plotly_chart(fig_heat, use_container_width=True, key="fig_heat")
else:
    st.info("Sin datos suficientes para el heatmap.")

# ── FILA 7: Antigüedad de tickets ──
st.divider()
st.subheader("⏳ Tickets agrupados por antigüedad")
if "Dias_Habiles" in df.columns and df["Dias_Habiles"].notna().any():
    bins_ant   = [0, 5, 15, 30, 60, float("inf")]
    labels_ant = ["🟢 0-5 días", "🟡 6-15 días", "🟠 16-30 días", "🔴 31-60 días", "⛔ +60 días"]
    df_ant = df.copy()
    df_ant["Antigüedad"] = pd.cut(
        df_ant["Dias_Habiles"], bins=bins_ant, labels=labels_ant, right=True
    )
    col_ant1, col_ant2 = st.columns([1, 2])

    with col_ant1:
        # Métricas por rango
        resumen_ant = df_ant["Antigüedad"].value_counts().reindex(labels_ant).fillna(0).astype(int)
        for rango, cantidad in resumen_ant.items():
            st.metric(str(rango), cantidad)

    with col_ant2:
        # Gráfica de barras por antigüedad y responsable
        df_ant_resp = (
            df_ant.groupby(["Antigüedad", "Responsable"])
                  .size().reset_index(name="Tickets")
        )
        df_ant_resp["Antigüedad"] = df_ant_resp["Antigüedad"].astype(str)
        fig_ant = px.bar(
            df_ant_resp,
            x="Antigüedad", y="Tickets", color="Responsable",
            color_discrete_sequence=PALETA,
            template="plotly_dark",
            category_orders={"Antigüedad": labels_ant},
            labels={"Antigüedad": "Rango de días", "Tickets": "Tickets"},
            barmode="stack",
        )
        fig_ant.update_layout(**LAYOUT)
        st.plotly_chart(fig_ant, use_container_width=True, key="fig_antiguedad")
else:
    st.info("Sin datos de días hábiles disponibles.")
